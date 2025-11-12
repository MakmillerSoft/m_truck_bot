"""
Генератор Excel файлів з даних БД
"""
import logging
from datetime import datetime
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.modules.database.manager import db_manager
from app.modules.admin.services.vehicle_management.shared.translations import translate_field_value

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Клас для експорту даних в Excel"""
    
    def __init__(self):
        self.wb = Workbook()
        # Видаляємо дефолтний лист
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
    
    def _style_header(self, ws, max_col: int):
        """Стилізувати заголовок таблиці"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _auto_size_columns(self, ws):
        """Автоматично підібрати ширину колонок"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    async def export_users(self) -> None:
        """Експортувати користувачів"""
        ws = self.wb.create_sheet("Користувачі")
        
        # Заголовки - ВСІ поля з БД
        headers = [
            "ID", "Telegram ID", "Ім'я", "Прізвище", "Username", "Телефон", 
            "Роль", "Активний", "Верифікований", "Дата реєстрації", "Дата оновлення"
        ]
        ws.append(headers)
        
        # Отримуємо користувачів
        users = await db_manager.get_all_users()
        logger.info(f"📊 Отримано {len(users)} користувачів з БД для експорту")
        
        for user in users:
            # Функція для безпечного перекладу
            def safe_translate(field_key: str, value: any) -> str:
                """Безпечно перекласти значення поля"""
                if not value or value == "":
                    return ""
                return translate_field_value(field_key, str(value))
            
            # Безпечне отримання з словника - ВСІ поля з БД + ПЕРЕКЛАДИ
            ws.append([
                user.get('id', ''),
                user.get('telegram_id', ''),
                user.get('first_name', '') or "",
                user.get('last_name', '') or "",
                user.get('username', '') or "",
                user.get('phone', '') or "",
                safe_translate('role', user.get('role')),  # ПЕРЕКЛАД
                "Так" if user.get('is_active') else "Ні",
                "Так" if user.get('is_verified') else "Ні",
                user.get('created_at', '') or "",
                user.get('updated_at', '') or ""
            ])
        
        self._style_header(ws, len(headers))
        self._auto_size_columns(ws)
        
        logger.info(f"✅ Експортовано {len(users)} користувачів")
    
    async def export_vehicles(self) -> None:
        """Експортувати авто"""
        ws = self.wb.create_sheet("Авто")
        
        # Заголовки - ВСІ поля з БД
        headers = [
            # Основна інформація
            "ID", "Тип", "Марка", "Модель", "VIN", "Рік", "Стан",
            # Ціна та валюта
            "Ціна", "Валюта", "Пробіг (км)",
            # Двигун
            "Об'єм двигуна (л)", "Потужність (к.с.)", "Тип палива",
            # Трансмісія та кузов
            "Коробка передач", "Тип кузова", "Радіус коліс",
            # Вантажні характеристики
            "Вантажопідйомність (кг)", "Загальна маса (кг)", "Габарити відсіку",
            # Локація та опис
            "Локація", "Опис",
            # Медіа
            "Кількість медіа", "Головне медіа", "Тип головного медіа", "Всі медіа (JSON)",
            # Статус та активність
            "Статус", "Активність",
            # Публікація
            "Опубліковано в групу", "Опубліковано в бот", "Дата публікації",
            "ID повідомлення в групі",
            # Дати
            "Дата зміни статусу", "Дата продажу",
            # Системні поля
            "Продавець ID", "Створено", "Оновлено"
        ]
        ws.append(headers)
        
        # Отримуємо авто
        vehicles = await db_manager.get_all_vehicles()
        logger.info(f"📊 Отримано {len(vehicles)} авто з БД для експорту")
        
        for vehicle in vehicles:
            # Обробка медіа (photos JSON - може містити фото та відео)
            photos_count = 0
            photos_json = ""
            if vehicle.get('photos'):
                try:
                    import json
                    photos_list = json.loads(vehicle.get('photos')) if isinstance(vehicle.get('photos'), str) else vehicle.get('photos')
                    photos_count = len(photos_list) if photos_list else 0
                    # Зберігаємо як JSON рядок для експорту
                    photos_json = json.dumps(photos_list, ensure_ascii=False) if photos_list else ""
                except:
                    photos_count = 0
                    photos_json = ""
            
            # Визначення типу головного медіа
            main_photo_type = ""
            main_photo_id = vehicle.get('main_photo', '') or ""
            if main_photo_id:
                if isinstance(main_photo_id, str) and main_photo_id.startswith("video:"):
                    main_photo_type = "Відео"
                else:
                    main_photo_type = "Фото"
            
            # Функція для безпечного перекладу
            def safe_translate(field_key: str, value: any) -> str:
                """Безпечно перекласти значення поля"""
                if not value or value == "":
                    return ""
                return translate_field_value(field_key, str(value))
            
            # Безпечне отримання з словника - ВСІ поля з ПЕРЕКЛАДАМИ
            ws.append([
                # Основна інформація
                vehicle.get('id', ''),
                safe_translate('vehicle_type', vehicle.get('vehicle_type')),  # ПЕРЕКЛАД
                vehicle.get('brand', '') or "",
                vehicle.get('model', '') or "",
                vehicle.get('vin_code', '') or "",
                vehicle.get('year', '') or "",
                safe_translate('condition', vehicle.get('condition')),  # ПЕРЕКЛАД
                # Ціна та валюта
                vehicle.get('price', '') or "",
                vehicle.get('currency', '') or "USD",
                vehicle.get('mileage', '') or "",
                # Двигун
                vehicle.get('engine_volume', '') or "",
                vehicle.get('power_hp', '') or "",
                safe_translate('fuel_type', vehicle.get('fuel_type')),  # ПЕРЕКЛАД
                # Трансмісія та кузов
                safe_translate('transmission', vehicle.get('transmission')),  # ПЕРЕКЛАД
                vehicle.get('body_type', '') or "",
                vehicle.get('wheel_radius', '') or "",
                # Вантажні характеристики
                vehicle.get('load_capacity', '') or "",
                vehicle.get('total_weight', '') or "",
                vehicle.get('cargo_dimensions', '') or "",
                # Локація та опис
                safe_translate('location', vehicle.get('location')),  # ПЕРЕКЛАД
                vehicle.get('description', '') or "",
                # Медіа
                photos_count,
                main_photo_id,
                main_photo_type,
                photos_json,
                # Статус та активність
                safe_translate('status', vehicle.get('status')),  # ПЕРЕКЛАД
                "Активне" if vehicle.get('is_active') else "Неактивне",
                # Публікація
                "Так" if vehicle.get('published_in_group') else "Ні",
                "Так" if vehicle.get('published_in_bot') else "Ні",
                vehicle.get('published_at', '') or "",
                vehicle.get('group_message_id', '') or "",
                # Дати
                vehicle.get('status_changed_at', '') or "",
                vehicle.get('sold_at', '') or "",
                # Системні поля
                vehicle.get('seller_id', '') or "",
                vehicle.get('created_at', '') or "",
                vehicle.get('updated_at', '') or ""
            ])
        
        self._style_header(ws, len(headers))
        self._auto_size_columns(ws)
        
        logger.info(f"✅ Експортовано {len(vehicles)} авто")
    
    async def export_requests(self) -> None:
        """Експортувати заявки"""
        ws = self.wb.create_sheet("Заявки")
        
        # Заголовки - ВСІ поля з БД
        headers = [
            "ID", "Користувач ID", "Авто ID", "Тип заявки", "Деталі", 
            "Статус", "Створено", "Оновлено"
        ]
        ws.append(headers)
        
        # Отримуємо заявки
        requests = await db_manager.get_all_requests()
        logger.info(f"📊 Отримано {len(requests)} заявок з БД для експорту")
        
        for request in requests:
            # Функція для безпечного перекладу
            def safe_translate(field_key: str, value: any) -> str:
                """Безпечно перекласти значення поля"""
                if not value or value == "":
                    return ""
                return translate_field_value(field_key, str(value))
            
            # request - це словник - тільки реальні поля з БД + ПЕРЕКЛАДИ
            ws.append([
                request.get('id', ''),
                request.get('user_id', ''),
                request.get('vehicle_id', ''),
                safe_translate('request_type', request.get('request_type')),  # ПЕРЕКЛАД
                request.get('details', ''),
                safe_translate('request_status', request.get('status')),  # ПЕРЕКЛАД
                request.get('created_at', ''),
                request.get('updated_at', '')
            ])
        
        self._style_header(ws, len(headers))
        self._auto_size_columns(ws)
        
        logger.info(f"✅ Експортовано {len(requests)} заявок")
    
    async def export_broadcasts(self) -> None:
        """Експортувати розсилки"""
        ws = self.wb.create_sheet("Розсилки")
        
        # Заголовки - ВСІ поля з БД
        headers = [
            "ID", "Текст", "Кнопка (текст)", "Кнопка (URL)", 
            "Тип медіа", "Media File ID", "Media Group ID", 
            "Статус", "Період повтору", "Заплановано", "Створено"
        ]
        ws.append(headers)
        
        # Отримуємо розсилки
        broadcasts = await db_manager.get_all_broadcasts_raw()
        logger.info(f"📊 Отримано {len(broadcasts)} розсилок з БД для експорту")
        
        for broadcast in broadcasts:
            # Функція для безпечного перекладу
            def safe_translate(field_key: str, value: any) -> str:
                """Безпечно перекласти значення поля"""
                if not value or value == "":
                    return ""
                return translate_field_value(field_key, str(value))
            
            # Безпечне отримання з словника - ВСІ поля з БД + ПЕРЕКЛАДИ
            text = broadcast.get('text', '') or ""
            text_short = (text[:50] + "...") if text and len(text) > 50 else text
            
            ws.append([
                broadcast.get('id', ''),
                text_short,
                broadcast.get('button_text', '') or "",
                broadcast.get('button_url', '') or "",
                safe_translate('media_type', broadcast.get('media_type')),  # ПЕРЕКЛАД
                broadcast.get('media_file_id', '') or "",
                broadcast.get('media_group_id', '') or "",
                safe_translate('broadcast_status', broadcast.get('status')),  # ПЕРЕКЛАД
                safe_translate('schedule_period', broadcast.get('schedule_period')),  # ПЕРЕКЛАД
                broadcast.get('scheduled_at', '') or "",
                broadcast.get('created_at', '') or ""
            ])
        
        self._style_header(ws, len(headers))
        self._auto_size_columns(ws)
        
        logger.info(f"✅ Експортовано {len(broadcasts)} розсилок")
    
    async def export_all(self) -> None:
        """Експортувати всі дані"""
        await self.export_users()
        await self.export_vehicles()
        await self.export_requests()
        await self.export_broadcasts()
        
        logger.info("✅ Експортовано всі дані")
    
    def save(self, filename: str) -> str:
        """Зберегти файл"""
        self.wb.save(filename)
        logger.info(f"📁 Файл збережено: {filename}")
        return filename


async def generate_excel_export(export_type: str) -> str:
    """
    Генерувати Excel файл з експортом даних
    
    Args:
        export_type: Тип експорту (users, vehicles, requests, broadcasts, all)
    
    Returns:
        str: Шлях до згенерованого файлу
    """
    exporter = ExcelExporter()
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"export_{export_type}_{timestamp}.xlsx"
    
    if export_type == "users":
        await exporter.export_users()
    elif export_type == "vehicles":
        await exporter.export_vehicles()
    elif export_type == "requests":
        await exporter.export_requests()
    elif export_type == "broadcasts":
        await exporter.export_broadcasts()
    elif export_type == "all":
        await exporter.export_all()
    else:
        raise ValueError(f"Невідомий тип експорту: {export_type}")
    
    return exporter.save(filename)
